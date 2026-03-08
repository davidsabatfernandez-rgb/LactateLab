from __future__ import annotations

import csv
import io
import json
from collections import defaultdict
from datetime import date, datetime, time
from typing import Any, Optional

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.athlete import Athlete, AthleteWeightHistory
from app.models.session import LactateSample, Session as AthleteSession, SessionInterval
from app.schemas.imports import ImportValidationIssue
from app.services.analytics import recalculate_athlete

COLUMN_DEFINITIONS = [
    {"key": "athlete", "label": "Atleta", "required": True, "aliases": ["athlete", "atleta", "name", "nombre"]},
    {"key": "date", "label": "Fecha", "required": True, "aliases": ["date", "fecha", "day"]},
    {"key": "time", "label": "Hora", "required": False, "aliases": ["time", "hora", "start_time"]},
    {"key": "discipline", "label": "Disciplina", "required": True, "aliases": ["discipline", "disciplina", "sport"]},
    {"key": "interval_duration", "label": "Duración intervalo", "required": True, "aliases": ["duration", "interval_duration", "duracion", "duración", "block_duration"]},
    {"key": "rest", "label": "Descanso", "required": False, "aliases": ["rest", "descanso", "recovery"]},
    {"key": "lactate", "label": "Lactato", "required": False, "aliases": ["lactate", "lactato", "mmol"]},
    {"key": "hr_avg", "label": "FC media", "required": False, "aliases": ["hr_avg", "fc_media", "avg_hr", "heart_rate_avg"]},
    {"key": "hr_max", "label": "FC máxima", "required": False, "aliases": ["hr_max", "fc_max", "max_hr", "heart_rate_max"]},
    {"key": "pace", "label": "Ritmo", "required": False, "aliases": ["pace", "ritmo", "pace_seconds_per_km"]},
    {"key": "power", "label": "Potencia", "required": False, "aliases": ["power", "potencia", "power_watts"]},
    {"key": "running_power", "label": "Potencia estimada corriendo", "required": False, "aliases": ["running_power", "potencia_estimada_corriendo", "running_power_watts"]},
    {"key": "weight", "label": "Peso", "required": False, "aliases": ["weight", "peso"]},
    {"key": "purpose", "label": "Propósito del intervalo", "required": False, "aliases": ["purpose", "proposito", "propósito", "interval_purpose"]},
    {"key": "session_type", "label": "Tipo de sesión", "required": False, "aliases": ["session_type", "tipo_sesion"]},
    {"key": "goal", "label": "Objetivo", "required": False, "aliases": ["goal", "objetivo"]},
    {"key": "sample_delay_seconds", "label": "Retraso muestra", "required": False, "aliases": ["sample_delay", "sample_delay_seconds", "delay", "retraso_muestra"]},
]


def _normalize_header(value: str) -> str:
    return "".join(character.lower() for character in value.strip() if character.isalnum() or character == "_")


def _suggest_mapping(headers: list[str]) -> dict[str, str]:
    normalized_headers = {_normalize_header(header): header for header in headers}
    mapping: dict[str, str] = {}
    for definition in COLUMN_DEFINITIONS:
        for alias in definition["aliases"]:
            match = normalized_headers.get(_normalize_header(alias))
            if match:
                mapping[definition["key"]] = match
                break
    return mapping


def _column_options() -> list[dict[str, Any]]:
    return [
        {"key": definition["key"], "label": definition["label"], "required": definition["required"], "aliases": definition["aliases"]}
        for definition in COLUMN_DEFINITIONS
    ]


def _parse_csv(file_bytes: bytes) -> tuple[list[str], list[dict[str, Any]]]:
    text = file_bytes.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    rows = [dict(row) for row in reader]
    return list(reader.fieldnames or []), rows


def _parse_excel(file_bytes: bytes) -> tuple[list[str], list[dict[str, Any]]]:
    workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    worksheet = workbook.active
    values = list(worksheet.iter_rows(values_only=True))
    if not values:
        return [], []
    headers = [str(cell).strip() if cell is not None else "" for cell in values[0]]
    rows = [{headers[index]: raw_row[index] if index < len(raw_row) else None for index in range(len(headers))} for raw_row in values[1:]]
    return headers, rows


def parse_tabular_file(filename: str, file_bytes: bytes) -> tuple[list[str], list[dict[str, Any]]]:
    lower_name = filename.lower()
    if lower_name.endswith(".csv"):
        return _parse_csv(file_bytes)
    if lower_name.endswith(".xlsx"):
        return _parse_excel(file_bytes)
    raise ValueError("Unsupported file type. Use CSV or XLSX.")


def _coerce_float(value: Any) -> Optional[float]:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", ".")
    return float(text) if text else None


def _coerce_int(value: Any) -> Optional[int]:
    parsed = _coerce_float(value)
    return int(round(parsed)) if parsed is not None else None


def _coerce_seconds(value: Any) -> Optional[int]:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        numeric = float(value)
        return int(round(numeric * 60)) if numeric <= 20 else int(round(numeric))
    text = str(value).strip()
    if ":" in text:
        parts = [int(part) for part in text.split(":")]
        if len(parts) == 2:
            return parts[0] * 60 + parts[1]
        if len(parts) == 3:
            return parts[0] * 3600 + parts[1] * 60 + parts[2]
    return int(round(float(text.replace(",", "."))))


def _coerce_pace_seconds(value: Any) -> Optional[float]:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        numeric = float(value)
        return numeric * 60 if numeric < 30 else numeric
    text = str(value).strip()
    if ":" in text:
        minutes, seconds = text.split(":")
        return float(int(minutes) * 60 + int(seconds))
    return float(text.replace(",", "."))


def _coerce_date(value: Any) -> Optional[date]:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _coerce_time(value: Any) -> Optional[time]:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.time().replace(microsecond=0)
    if isinstance(value, time):
        return value.replace(microsecond=0)
    text = str(value).strip()
    for fmt in ("%H:%M", "%H:%M:%S"):
        try:
            return datetime.strptime(text, fmt).time()
        except ValueError:
            continue
    return None


def parse_json_form_field(raw_value: Optional[str]) -> dict[str, str]:
    if not raw_value:
        return {}
    parsed = json.loads(raw_value)
    if not isinstance(parsed, dict):
        raise ValueError("Expected JSON object")
    return {str(key): "" if value is None else str(value) for key, value in parsed.items()}


def _normalize_row(raw_row: dict[str, Any], mapping: dict[str, str], defaults: dict[str, str]) -> dict[str, Any]:
    def value_for(key: str) -> Any:
        column = mapping.get(key)
        raw_value = raw_row.get(column) if column else None
        return defaults.get(key) if raw_value in (None, "") else raw_value

    return {
        "athlete": str(value_for("athlete")).strip() if value_for("athlete") not in (None, "") else None,
        "date": _coerce_date(value_for("date")),
        "time": _coerce_time(value_for("time")) or time(6, 0),
        "discipline": str(value_for("discipline")).strip().lower() if value_for("discipline") not in (None, "") else None,
        "interval_duration": _coerce_seconds(value_for("interval_duration")),
        "rest": _coerce_seconds(value_for("rest")),
        "lactate": _coerce_float(value_for("lactate")),
        "hr_avg": _coerce_int(value_for("hr_avg")),
        "hr_max": _coerce_int(value_for("hr_max")),
        "pace": _coerce_pace_seconds(value_for("pace")),
        "power": _coerce_float(value_for("power")),
        "running_power": _coerce_float(value_for("running_power")),
        "weight": _coerce_float(value_for("weight")),
        "purpose": str(value_for("purpose")).strip() if value_for("purpose") not in (None, "") else "intervalos",
        "session_type": str(value_for("session_type")).strip() if value_for("session_type") not in (None, "") else "intervalos",
        "goal": str(value_for("goal")).strip() if value_for("goal") not in (None, "") else "Importación masiva",
        "sample_delay_seconds": _coerce_seconds(value_for("sample_delay_seconds")) or 0,
    }


def _validate_normalized_row(row_number: int, normalized: dict[str, Any], athletes_by_name: dict[str, Athlete]) -> list[str]:
    errors = []
    if not normalized["athlete"]:
        errors.append("Falta atleta.")
    elif normalized["athlete"].lower() not in athletes_by_name:
        errors.append(f"Atleta no encontrado: {normalized['athlete']}.")
    if not normalized["date"]:
        errors.append("Fecha inválida o ausente.")
    if not normalized["discipline"]:
        errors.append("Disciplina ausente.")
    if not normalized["interval_duration"] or normalized["interval_duration"] <= 0:
        errors.append("Duración de intervalo inválida.")
    if normalized["lactate"] is not None and normalized["lactate"] <= 0:
        errors.append("Lactato debe ser > 0.")
    return errors


def build_import_preview(
    db: Session,
    filename: str,
    file_bytes: bytes,
    mapping: Optional[dict[str, str]] = None,
    defaults: Optional[dict[str, str]] = None,
) -> dict[str, Any]:
    headers, rows = parse_tabular_file(filename, file_bytes)
    suggested_mapping = _suggest_mapping(headers)
    active_mapping = dict(suggested_mapping)
    if mapping:
        active_mapping.update(mapping)
    defaults = defaults or {}

    athletes = db.scalars(select(Athlete)).all()
    athletes_by_name = {athlete.name.lower(): athlete for athlete in athletes}
    missing_required = [definition["key"] for definition in COLUMN_DEFINITIONS if definition["required"] and not active_mapping.get(definition["key"])]

    issues: list[ImportValidationIssue] = []
    for item in missing_required:
        issues.append(ImportValidationIssue(row_number=None, message=f"Falta mapear columna requerida: {item}.", severity="error"))

    preview_rows = []
    for index, raw_row in enumerate(rows[:12], start=2):
        normalized = _normalize_row(raw_row, active_mapping, defaults)
        row_errors = _validate_normalized_row(index, normalized, athletes_by_name) if not missing_required else []
        preview_rows.append({"row_number": index, "values": raw_row, "normalized": normalized, "errors": row_errors})
        for error in row_errors:
            issues.append(ImportValidationIssue(row_number=index, message=error, severity="error"))

    return {
        "filename": filename,
        "headers": headers,
        "column_options": _column_options(),
        "suggested_mapping": suggested_mapping,
        "active_mapping": active_mapping,
        "defaults": defaults,
        "missing_required_fields": missing_required,
        "issues": issues,
        "preview_rows": preview_rows,
        "total_rows": len(rows),
        "can_import": not any(issue.severity == "error" for issue in issues),
    }


def commit_import(db: Session, filename: str, file_bytes: bytes, mapping: dict[str, str], defaults: Optional[dict[str, str]] = None) -> dict[str, Any]:
    preview = build_import_preview(db, filename, file_bytes, mapping=mapping, defaults=defaults or {})
    if not preview["can_import"]:
        return {"imported_sessions": 0, "imported_intervals": 0, "affected_athletes": [], "issues": preview["issues"]}

    _, rows = parse_tabular_file(filename, file_bytes)
    athletes = db.scalars(select(Athlete)).all()
    athletes_by_name = {athlete.name.lower(): athlete for athlete in athletes}

    grouped_rows: dict[tuple[int, datetime, str, str, str], list[dict[str, Any]]] = defaultdict(list)
    issues: list[ImportValidationIssue] = []
    for index, raw_row in enumerate(rows, start=2):
        normalized = _normalize_row(raw_row, mapping, defaults or {})
        row_errors = _validate_normalized_row(index, normalized, athletes_by_name)
        if row_errors:
            for error in row_errors:
                issues.append(ImportValidationIssue(row_number=index, message=error, severity="error"))
            continue
        athlete = athletes_by_name[normalized["athlete"].lower()]
        session_dt = datetime.combine(normalized["date"], normalized["time"])
        grouped_rows[(athlete.id, session_dt, normalized["discipline"], normalized["session_type"], normalized["goal"])].append(normalized)

    if issues:
        return {"imported_sessions": 0, "imported_intervals": 0, "affected_athletes": [], "issues": issues}

    imported_sessions = 0
    imported_intervals = 0
    affected_athletes: set[int] = set()
    for group_key, normalized_rows in grouped_rows.items():
        athlete_id, session_dt, discipline, session_type, goal = group_key
        session = AthleteSession(
            athlete_id=athlete_id,
            performed_at=session_dt,
            discipline=discipline,
            session_type=session_type,
            goal=goal,
            comments=f"Importado desde {filename}",
        )
        for order_index, normalized in enumerate(normalized_rows, start=1):
            interval = SessionInterval(
                order_index=order_index,
                duration_seconds=normalized["interval_duration"],
                rest_seconds=normalized["rest"],
                rest_type="imported",
                heart_rate_avg=normalized["hr_avg"],
                heart_rate_max=normalized["hr_max"],
                pace_seconds_per_km=normalized["pace"],
                power_watts=normalized["power"],
                running_power_watts=normalized["running_power"],
                purpose=normalized["purpose"],
                notes="Intervalo importado",
            )
            if normalized["lactate"] is not None:
                interval.lactate_sample = LactateSample(
                    lactate_mmol=normalized["lactate"],
                    sample_delay_seconds=normalized["sample_delay_seconds"],
                    sample_timing_label="importado",
                    sampling_notes="Importación masiva",
                )
            session.intervals.append(interval)
            imported_intervals += 1
        db.add(session)
        imported_sessions += 1
        affected_athletes.add(athlete_id)
        last_weight = normalized_rows[-1].get("weight")
        if last_weight is not None:
            athlete = next(item for item in athletes if item.id == athlete_id)
            athlete.weight = last_weight
            db.add(AthleteWeightHistory(athlete_id=athlete_id, recorded_at=session_dt.date(), weight=last_weight, source="bulk_import"))

    db.commit()
    for athlete_id in sorted(affected_athletes):
        recalculate_athlete(db, athlete_id)

    return {"imported_sessions": imported_sessions, "imported_intervals": imported_intervals, "affected_athletes": sorted(affected_athletes), "issues": issues}
